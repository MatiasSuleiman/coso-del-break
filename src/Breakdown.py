from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

try:
    from src.buscador_adapter import normalizar_datetime_naive
except ModuleNotFoundError:
    from buscador_adapter import normalizar_datetime_naive

class Breakdown:
    @classmethod
    def con_mails_manejado_por(self, mails, abogado, path, sistema):
        return self(mails, abogado, path=path, sistema=sistema)
    
    def __init__(self, mails, abogado, path, sistema):
        self.workbook = Workbook()
        self.excel = self.workbook.active
        self.initialize_excel_file(self.ordenar_por_fecha(mails), abogado, path=path, sistema=sistema)

    def ordenar_por_fecha(self, mails):
        return sorted(mails, key=lambda mail: normalizar_datetime_naive(mail.date))

    def autoajustar_columnas(self):
        for indice, columna in enumerate(self.excel.iter_cols(), start=1):
            maximo = 0
            for casilla in columna:
                valor = "" if casilla.value is None else str(casilla.value)
                maximo = max(maximo, len(valor))
            self.excel.column_dimensions[get_column_letter(indice)].width = maximo + 2

    def initialize_excel_file(self, mails, abogado, path, sistema):
        self.excel['A1'] = "Date"
        self.excel['B1'] = "Description"
        self.excel['C1'] = "Minutes"
        self.excel['D1'] = "Hours"

        self.excel['A1'].font = Font(size=14)
        self.excel['B1'].font = Font(size=14)
        self.excel['C1'].font = Font(size=14)
        self.excel['D1'].font = Font(size=14)

        for i in range(len(mails)):
            mail = mails[i]
            self.excel[f"A{i+2}"] = mail.date.strftime("%d/%m/%Y")
            self.excel[f"B{i+2}"] = sistema.ver_descripcion_de(mail)

        for i in range(len(mails)):
           mail = mails[i]
           self.excel[f"C{i+2}"] = sistema.ver_minutos_de(mail)
           self.excel[f"D{i+2}"] = f"=C{i+2} / 60"

        self.autoajustar_columnas()
        self.workbook.save(path)


    def tiene_en(self, casilla, string):
        return self.excel[casilla].value == string


    def casilla_tiene_fuente_tamanio(self, casilla, tamanio):
        return self.excel[casilla].font.size == tamanio
